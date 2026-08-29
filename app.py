import streamlit as st
import openpyxl
import os
from datetime import datetime

st.set_page_config(page_title="Zomato Tracker", page_icon="📊")
st.title("Zomato Smart Tracker v3.0")

# User Input Box
amount = st.number_input("Aaj ki kamai (Rs.) yahan likhein:", min_value=0, step=100)

if st.button("Excel Me Auto-Save Karein"):
    if amount == 0:
        st.warning("⚠️ Kripya pehle sahi amount likhein!")
    else:
        file_path = "ZomatoSmartTracker.xlsx"
        
        now = datetime.now()
        current_date = now.strftime("%Y-%m-%d")
        current_time = now.strftime("%H:%M:%S")
        
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Smart_Earnings"
            sheet["A1"], sheet["B1"], sheet["C1"] = "Date", "Time", "Earnings (Rs.)"
            
        next_row = sheet.max_row + 1
        sheet[f"A{next_row}"] = current_date
        sheet[f"B{next_row}"] = current_time
        sheet[f"C{next_row}"] = int(amount)
        
        wb.save(file_path)
        st.success(f"✅ Rs. {amount} save ho gaye! \n Date: {current_date} | Time: {current_time}")
        
        with open(file_path, "rb") as file:
            st.download_button(
                label="📥 Apni Excel File Phone Me Download Karein",
                data=file,
                file_name="ZomatoSmartTracker.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
           
            )
